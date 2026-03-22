"""
Input validation & API key verification.
"""

import os
from pathlib import Path

from dotenv import load_dotenv

load_dotenv()


def validate_excel_file(file_path: str) -> dict:
    """
    Validate an uploaded Excel file.
    Returns {"valid": True} or {"valid": False, "error": "..."}.
    """
    p = Path(file_path)

    if not p.exists():
        return {"valid": False, "error": f"File not found: {file_path}"}

    if p.suffix.lower() not in (".xlsx", ".xlsm"):
        return {"valid": False, "error": f"Invalid file type: {p.suffix}. Expected .xlsx or .xlsm"}

    if p.stat().st_size == 0:
        return {"valid": False, "error": "File is empty"}

    if p.stat().st_size > 50 * 1024 * 1024:  # 50MB
        return {"valid": False, "error": "File is too large (max 50MB)"}

    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True)
        sheet_names = wb.sheetnames

        # Check for Raw Data sheet
        raw_data_sheets = [s for s in sheet_names if "raw data" in s.lower()]
        if not raw_data_sheets:
            return {
                "valid": False,
                "error": f"No 'Raw Data' sheet found. Available sheets: {', '.join(sheet_names)}",
            }

        wb.close()
        return {"valid": True, "sheets": sheet_names}

    except Exception as e:
        return {"valid": False, "error": f"Cannot read Excel file: {e}"}


def validate_api_keys(model: str = "gemini-2.5-pro") -> dict:
    """
    Check if the required API key is set for the given model.
    Returns {"valid": True, "provider": "..."} or {"valid": False, "error": "..."}.
    """
    from image_analyzer import MODEL_REGISTRY, get_provider

    if model in MODEL_REGISTRY:
        provider = MODEL_REGISTRY[model][0]
    else:
        provider = get_provider(model)

    key_env_map = {
        "claude": ("ANTHROPIC_API_KEY", "https://console.anthropic.com/"),
        "gemini": ("GEMINI_API_KEY", "https://aistudio.google.com/apikey"),
        "fireworks": ("FIREWORKS_API_KEY", "https://fireworks.ai/"),
    }

    env_var, url = key_env_map.get(provider, ("UNKNOWN", ""))
    api_key = os.getenv(env_var, "")

    if not api_key:
        return {
            "valid": False,
            "provider": provider,
            "error": f"{env_var} is not set. Get your key at {url}",
        }

    # Basic format check
    if provider == "claude" and not api_key.startswith("sk-ant-"):
        return {
            "valid": False,
            "provider": provider,
            "error": f"{env_var} doesn't look like a valid Anthropic key (should start with 'sk-ant-')",
        }

    if provider == "gemini" and not api_key.startswith("AIza"):
        return {
            "valid": False,
            "provider": provider,
            "error": f"{env_var} doesn't look like a valid Google AI key (should start with 'AIza')",
        }

    return {"valid": True, "provider": provider, "env_var": env_var}


def validate_image_url(url: str) -> bool:
    """Quick check if a string looks like a valid image URL."""
    if not url or not isinstance(url, str):
        return False
    url = url.strip()
    return url.startswith("http://") or url.startswith("https://")


def estimate_processing_time(
    num_images: int,
    model: str = "gemini-2.5-pro",
    enhancements_enabled: bool = True,
) -> dict:
    """
    Estimate processing time based on model RPM and image count.
    Accounts for real-world API latency (~5s per call) and network overhead.
    """
    from rate_limiter import RateLimiter

    limiter = RateLimiter(model)
    safe_workers = limiter.get_safe_workers()

    # Each image requires API calls:
    # - 1 main analysis call (~5s latency)
    # - 1 OCR call (~3s) if enhancements enabled
    # - 1 SKU refinement call (~4s) if enhancements enabled
    calls_per_image = 1
    avg_call_latency = 5.0  # seconds per API call (real-world average)

    if enhancements_enabled:
        calls_per_image = 3  # main + OCR + SKU refine

    total_api_calls = num_images * calls_per_image

    # Rate limit: how fast can we START calls?
    # With RPM limit, we can start (RPM / 60) calls per second
    calls_per_second = limiter.rpm / 60.0

    # Time to issue all calls (rate-limited)
    time_to_issue = total_api_calls / calls_per_second

    # Time for last batch of calls to complete
    # With N workers, last batch takes avg_call_latency seconds
    completion_overhead = avg_call_latency

    # Image fetch overhead (~2s per image, parallelized)
    fetch_time = (num_images * 2.0) / safe_workers

    total_seconds = time_to_issue + completion_overhead + fetch_time

    # Add 20% buffer for network variance
    total_seconds *= 1.2

    minutes = total_seconds / 60

    return {
        "estimated_seconds": int(total_seconds),
        "estimated_minutes": round(minutes, 1),
        "display": f"~{int(minutes)}min" if minutes >= 1 else f"~{int(total_seconds)}sec",
        "workers": safe_workers,
        "rpm": limiter.rpm,
        "per_image_seconds": round(total_seconds / num_images, 1) if num_images > 0 else 0,
    }
