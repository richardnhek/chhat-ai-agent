"""
Integration tests for the CHHAT Cigarette Brand Analyzer pipeline.
Tests the full flow with mocked API responses.
"""

import io
import json
import os
import threading
import time
from pathlib import Path
from unittest.mock import patch, MagicMock, call

import pytest
from PIL import Image as PILImage
from openpyxl import load_workbook


# ── Fixtures ──────────────────────────────────────────────────────────────


@pytest.fixture
def test_jpeg_bytes():
    """Create a small 100x100 JPEG image programmatically."""
    img = PILImage.new("RGB", (100, 100), color=(200, 50, 50))
    buf = io.BytesIO()
    img.save(buf, format="JPEG", quality=85)
    return buf.getvalue()


@pytest.fixture
def test_jpeg_large():
    """Create a larger JPEG for enhancement testing (adds more texture)."""
    import numpy as np
    # Create an image with some noise/texture so blur detection has something to work with
    arr = np.random.randint(0, 255, (200, 200, 3), dtype=np.uint8)
    img = PILImage.fromarray(arr)
    buf = io.BytesIO()
    img.save(buf, format="JPEG", quality=90)
    return buf.getvalue()


@pytest.fixture
def mock_analysis_result():
    """Standard AI analysis result."""
    return {
        "brands_found": [
            {"brand": "MEVIUS", "skus": ["MEVIUS ORIGINAL"], "notes": "Blue pack, center shelf"},
            {"brand": "ARA", "skus": ["ARA RED"], "notes": "Red pack, left side"},
        ],
        "brand_count": 2,
        "unidentified_packs": 0,
        "confidence": "high",
        "notes": "Clear image, good visibility",
    }


@pytest.fixture
def mock_results():
    """List of processed outlet results for output building."""
    return [
        {
            "serial": "S001",
            "brands": ["MEVIUS", "ARA"],
            "skus": ["MEVIUS ORIGINAL", "ARA RED"],
            "thumbnails": [],
            "unidentified_packs": 0,
            "confidence": "high",
            "confidence_score": 92,
            "confidence_factors": {},
            "error": None,
        },
        {
            "serial": "S002",
            "brands": ["555"],
            "skus": ["555 GOLD"],
            "thumbnails": [],
            "unidentified_packs": 1,
            "confidence": "medium",
            "confidence_score": 68,
            "confidence_factors": {},
            "error": None,
        },
    ]


# ═══════════════════════════════════════════════════════════════════════════════
# 1. test_fetch_image_caching
# ═══════════════════════════════════════════════════════════════════════════════


class TestFetchImageCaching:
    def test_fetch_image_caching(self, test_jpeg_bytes, tmp_path, monkeypatch):
        """Verify cache stores images on first fetch and returns them on second fetch."""
        import image_cache
        monkeypatch.setattr(image_cache, "CACHE_DIR", tmp_path / "cache")

        url = "https://ifieldTH.ipsos.com/getImage.asp?ID=12345"

        # Mock requests.get to return our test image
        mock_response = MagicMock()
        mock_response.content = test_jpeg_bytes
        mock_response.headers = {"Content-Type": "image/jpeg"}
        mock_response.raise_for_status = MagicMock()

        with patch("image_analyzer.requests.get", return_value=mock_response) as mock_get:
            from image_analyzer import fetch_image

            # First call: should hit network
            data1, media1 = fetch_image(url)
            assert mock_get.call_count == 1
            assert media1 == "image/jpeg"
            assert len(data1) > 0

            # Second call: should come from cache, no additional network call
            data2, media2 = fetch_image(url)
            assert mock_get.call_count == 1  # still 1, no new request
            assert media2 == "image/jpeg"
            # Data should be the same (both went through _resize_image)
            assert data1 == data2

    def test_fetch_image_cache_disabled(self, test_jpeg_bytes, tmp_path, monkeypatch):
        """Verify use_cache=False bypasses the cache."""
        import image_cache
        monkeypatch.setattr(image_cache, "CACHE_DIR", tmp_path / "cache")

        url = "https://ifieldTH.ipsos.com/getImage.asp?ID=99999"

        mock_response = MagicMock()
        mock_response.content = test_jpeg_bytes
        mock_response.headers = {"Content-Type": "image/jpeg"}
        mock_response.raise_for_status = MagicMock()

        with patch("image_analyzer.requests.get", return_value=mock_response) as mock_get:
            from image_analyzer import fetch_image

            fetch_image(url, use_cache=False)
            fetch_image(url, use_cache=False)
            # Both calls should hit network
            assert mock_get.call_count == 2

    def test_cache_stats(self, test_jpeg_bytes, tmp_path, monkeypatch):
        """Verify get_cache_stats returns correct counts."""
        import image_cache
        cache_dir = tmp_path / "cache"
        monkeypatch.setattr(image_cache, "CACHE_DIR", cache_dir)

        stats = image_cache.get_cache_stats()
        assert stats["total_images"] == 0

        image_cache.cache_image("https://example.com/1.jpg", test_jpeg_bytes, "image/jpeg")
        image_cache.cache_image("https://example.com/2.jpg", test_jpeg_bytes, "image/jpeg")

        stats = image_cache.get_cache_stats()
        assert stats["total_images"] == 2
        # Cache size should be positive (images are small so check raw bytes via file)
        total_bytes = sum(p.stat().st_size for p in cache_dir.iterdir() if p.suffix == ".img")
        assert total_bytes > 0

    def test_clear_cache(self, test_jpeg_bytes, tmp_path, monkeypatch):
        """Verify clear_cache removes all cached files."""
        import image_cache
        monkeypatch.setattr(image_cache, "CACHE_DIR", tmp_path / "cache")

        image_cache.cache_image("https://example.com/1.jpg", test_jpeg_bytes, "image/jpeg")
        assert image_cache.get_cache_stats()["total_images"] == 1

        image_cache.clear_cache()
        assert image_cache.get_cache_stats()["total_images"] == 0

    def test_cache_eviction(self, tmp_path, monkeypatch):
        """Verify LRU eviction when cache exceeds max size."""
        import image_cache
        monkeypatch.setattr(image_cache, "CACHE_DIR", tmp_path / "cache")
        # Set a tiny max cache size (2KB) to trigger eviction
        monkeypatch.setattr(image_cache, "MAX_CACHE_SIZE_BYTES", 2048)

        # Create images that together exceed 2KB
        big_data = b"\x00" * 1500  # 1.5KB each

        # Write first image file directly (to control mtime)
        image_cache.cache_image("https://example.com/old.jpg", big_data, "image/jpeg")
        # Backdate the old entry
        old_key = image_cache._url_to_key("https://example.com/old.jpg")
        old_path = image_cache._data_path(old_key)
        os.utime(old_path, (1000, 1000))

        # Write second image — should trigger eviction of old.jpg
        image_cache.cache_image("https://example.com/new.jpg", big_data, "image/jpeg")

        # Old entry should have been evicted
        assert image_cache.get_cached_image("https://example.com/old.jpg") is None
        # New entry should still be present
        assert image_cache.get_cached_image("https://example.com/new.jpg") is not None


# ═══════════════════════════════════════════════════════════════════════════════
# 2. test_enhance_image_pipeline
# ═══════════════════════════════════════════════════════════════════════════════


class TestEnhanceImagePipeline:
    def test_enhance_image_produces_valid_jpeg(self, test_jpeg_large):
        """Run enhance_image and verify output is a valid JPEG."""
        from enhancements import enhance_image

        enhanced = enhance_image(test_jpeg_large)
        assert len(enhanced) > 0
        # Should be loadable as an image
        img = PILImage.open(io.BytesIO(enhanced))
        assert img.format == "JPEG"

    def test_blur_score_changes_after_enhancement(self, test_jpeg_large):
        """Verify blur score is affected by enhancement."""
        from enhancements import enhance_image, compute_blur_score

        original_score = compute_blur_score(test_jpeg_large)
        enhanced = enhance_image(test_jpeg_large)
        enhanced_score = compute_blur_score(enhanced)

        # Both scores should be positive (valid images)
        assert original_score > 0
        assert enhanced_score > 0
        # Scores should differ (enhancement changes the image)
        assert original_score != enhanced_score

    def test_enhance_image_handles_invalid_data(self):
        """Enhancement should return original data if image can't be decoded."""
        from enhancements import enhance_image

        bad_data = b"not an image"
        result = enhance_image(bad_data)
        assert result == bad_data

    def test_blur_score_zero_for_invalid_image(self):
        """Blur score should be 0 for invalid image data."""
        from enhancements import compute_blur_score

        score = compute_blur_score(b"not an image")
        assert score == 0.0


# ═══════════════════════════════════════════════════════════════════════════════
# 3. test_parse_response_real_formats
# ═══════════════════════════════════════════════════════════════════════════════


class TestParseResponseRealFormats:
    def test_gemini_markdown_wrapped_response(self):
        """Gemini often wraps JSON in ```json ... ``` blocks."""
        from image_analyzer import _parse_response

        gemini_response = '```json\n{\n    "brands_found": [\n        {"brand": "MEVIUS", "skus": ["MEVIUS ORIGINAL"], "notes": "blue pack center"}\n    ],\n    "brand_count": 1,\n    "unidentified_packs": 0,\n    "confidence": "high",\n    "notes": "Clear display case"\n}\n```'
        result = _parse_response(gemini_response)
        assert result["brand_count"] == 1
        assert result["brands_found"][0]["brand"] == "MEVIUS"
        assert result["confidence"] == "high"

    def test_claude_raw_json_response(self):
        """Claude typically returns raw JSON without markdown wrapping."""
        from image_analyzer import _parse_response

        claude_response = '{"brands_found": [{"brand": "ARA", "skus": ["ARA RED"], "notes": "red pack left"}, {"brand": "555", "skus": ["555 GOLD"], "notes": "gold pack right"}], "brand_count": 2, "unidentified_packs": 1, "confidence": "medium", "notes": "Some glare on glass"}'
        result = _parse_response(claude_response)
        assert result["brand_count"] == 2
        assert len(result["brands_found"]) == 2
        brands = {b["brand"] for b in result["brands_found"]}
        assert brands == {"ARA", "555"}

    def test_response_with_extra_text(self):
        """Sometimes models prefix or suffix with commentary."""
        from image_analyzer import _parse_response

        response = 'Here is my analysis of the image:\n\n{"brands_found": [{"brand": "ESSE", "skus": ["ESSE CHANGE"]}], "brand_count": 1, "unidentified_packs": 0, "confidence": "high", "notes": ""}\n\nI hope this helps!'
        result = _parse_response(response)
        assert result["brand_count"] == 1
        assert result["brands_found"][0]["brand"] == "ESSE"

    def test_empty_result_response(self):
        """No cigarettes found response."""
        from image_analyzer import _parse_response

        response = '{"brands_found": [], "brand_count": 0, "unidentified_packs": 0, "confidence": "high", "notes": "Image shows a drinks display, no cigarettes visible"}'
        result = _parse_response(response)
        assert result["brand_count"] == 0
        assert result["brands_found"] == []
        assert result["confidence"] == "high"

    def test_completely_unparseable(self):
        """Totally broken response should return safe fallback."""
        from image_analyzer import _parse_response

        result = _parse_response("I cannot analyze this image due to content policy.")
        assert result["brands_found"] == []
        assert result["confidence"] == "low"
        assert "Could not parse" in result["notes"]


# ═══════════════════════════════════════════════════════════════════════════════
# 4. test_analyze_image_enhanced_mocked
# ═══════════════════════════════════════════════════════════════════════════════


class TestAnalyzeImageEnhancedMocked:
    def test_full_pipeline_calls_in_order(self, test_jpeg_bytes, mock_analysis_result):
        """Verify enhance -> OCR -> analyze -> SKU refine are called in the right order."""
        from enhancements import analyze_image_enhanced

        call_order = []

        def mock_enhance(data):
            call_order.append("enhance")
            return data

        def mock_ocr(data, media, model, keys):
            call_order.append("ocr")
            return ["MEVIUS", "ARA"]

        def mock_analyze(data, media, model=None, api_keys=None, correction_context=""):
            call_order.append("analyze")
            return mock_analysis_result

        def mock_refine(data, media, brands, model, keys):
            call_order.append("sku_refine")
            return [
                {"brand": "MEVIUS", "sku": "MEVIUS ORIGINAL", "reasoning": "blue pack"},
                {"brand": "ARA", "sku": "ARA RED", "reasoning": "red pack"},
            ]

        with patch("enhancements.enhance_image", side_effect=mock_enhance), \
             patch("enhancements.extract_text_ocr", side_effect=mock_ocr), \
             patch("image_analyzer.analyze_image", side_effect=mock_analyze), \
             patch("enhancements.refine_skus", side_effect=mock_refine):

            result = analyze_image_enhanced(
                test_jpeg_bytes,
                "image/jpeg",
                model="gemini-2.5-pro",
                api_keys={"gemini": "fake-key"},
                enable_enhancement=True,
                enable_ocr=True,
                enable_sku_refinement=True,
            )

        assert call_order == ["enhance", "ocr", "analyze", "sku_refine"]
        assert result["brand_count"] == 2
        assert "blur_score" in result

    def test_pipeline_skips_disabled_steps(self, test_jpeg_bytes, mock_analysis_result):
        """Verify disabled steps are not called."""
        from enhancements import analyze_image_enhanced

        call_order = []

        def mock_analyze(data, media, model=None, api_keys=None, correction_context=""):
            call_order.append("analyze")
            return mock_analysis_result

        with patch("enhancements.enhance_image") as mock_enh, \
             patch("enhancements.extract_text_ocr") as mock_ocr, \
             patch("image_analyzer.analyze_image", side_effect=mock_analyze), \
             patch("enhancements.refine_skus") as mock_refine:

            result = analyze_image_enhanced(
                test_jpeg_bytes,
                "image/jpeg",
                model="gemini-2.5-pro",
                api_keys={"gemini": "fake-key"},
                enable_enhancement=False,
                enable_ocr=False,
                enable_sku_refinement=False,
            )

        mock_enh.assert_not_called()
        mock_ocr.assert_not_called()
        mock_refine.assert_not_called()
        assert "analyze" in call_order

    def test_pipeline_handles_api_error(self, test_jpeg_bytes):
        """Pipeline should propagate error from analyze_image."""
        from enhancements import analyze_image_enhanced

        error_result = {"error": "gemini API error: quota exceeded"}

        with patch("enhancements.enhance_image", side_effect=lambda d: d), \
             patch("enhancements.extract_text_ocr", return_value=[]), \
             patch("image_analyzer.analyze_image", return_value=error_result), \
             patch("enhancements.refine_skus") as mock_refine:

            result = analyze_image_enhanced(
                test_jpeg_bytes,
                "image/jpeg",
                model="gemini-2.5-pro",
                api_keys={"gemini": "fake-key"},
            )

        assert "error" in result
        # SKU refinement should NOT be called when analysis errors
        mock_refine.assert_not_called()


# ═══════════════════════════════════════════════════════════════════════════════
# 5. test_build_output_creates_valid_excel
# ═══════════════════════════════════════════════════════════════════════════════


class TestBuildOutput:
    def test_creates_valid_excel_with_correct_columns(self, mock_results, tmp_path):
        """Verify build_output creates a valid Excel with expected headers."""
        from process import build_output

        output_path = str(tmp_path / "test_output.xlsx")
        build_output(mock_results, output_path)

        assert Path(output_path).exists()
        wb = load_workbook(output_path)
        ws = wb.active
        assert ws.title == "Results"

        # Check headers
        expected_headers = [
            "Serial Number", "Q32 Image 1", "Q32 Image 2", "Q32 Image 3",
            "Q12A - Brands", "Q12B - SKUs", "Brand Count",
            "Unidentified Packs", "Confidence", "Status",
        ]
        for i, header in enumerate(expected_headers, 1):
            assert ws.cell(row=1, column=i).value == header

    def test_data_rows_match_results(self, mock_results, tmp_path):
        """Verify data rows contain the right serial numbers and brand counts."""
        from process import build_output

        output_path = str(tmp_path / "test_output.xlsx")
        build_output(mock_results, output_path)

        wb = load_workbook(output_path)
        ws = wb.active

        # Row 2 = first result
        assert ws.cell(row=2, column=1).value == "S001"
        assert ws.cell(row=2, column=7).value == 2  # brand count
        assert ws.cell(row=2, column=10).value == "OK"

        # Row 3 = second result
        assert ws.cell(row=3, column=1).value == "S002"
        assert ws.cell(row=3, column=7).value == 1
        assert ws.cell(row=3, column=8).value == 1  # unidentified packs

    def test_error_result_shows_error_status(self, tmp_path):
        """Verify an error result gets STATUS=ERROR."""
        from process import build_output

        results = [{
            "serial": "S999",
            "brands": [],
            "skus": [],
            "thumbnails": [],
            "unidentified_packs": 0,
            "confidence": "low",
            "confidence_score": 0,
            "confidence_factors": {},
            "error": "Timeout fetching image",
        }]
        output_path = str(tmp_path / "error_output.xlsx")
        build_output(results, output_path)

        wb = load_workbook(output_path)
        ws = wb.active
        assert ws.cell(row=2, column=10).value == "ERROR"


# ═══════════════════════════════════════════════════════════════════════════════
# 6. test_build_client_format_matches_spec
# ═══════════════════════════════════════════════════════════════════════════════


class TestBuildClientFormat:
    def test_headers_at_correct_rows(self, mock_results, tmp_path):
        """Client format has field codes at row 5, descriptions at row 6, data from row 7."""
        from process import build_client_format

        output_path = str(tmp_path / "client_output.xlsx")
        build_client_format(mock_results, output_path)

        wb = load_workbook(output_path)
        ws = wb.active

        # Row 5: field codes
        assert ws.cell(row=5, column=1).value == "Respondent.Serial"
        assert ws.cell(row=5, column=2).value == "Q12A"
        assert ws.cell(row=5, column=3).value == "Q12B"

        # Row 6: descriptions
        assert ws.cell(row=6, column=1).value == "Serial number"
        assert "BRAND" in ws.cell(row=6, column=2).value.upper()
        assert "SKU" in ws.cell(row=6, column=3).value.upper()

    def test_data_starts_at_row_7(self, mock_results, tmp_path):
        """Verify data rows start at row 7."""
        from process import build_client_format

        output_path = str(tmp_path / "client_output.xlsx")
        build_client_format(mock_results, output_path)

        wb = load_workbook(output_path)
        ws = wb.active

        # Row 7 = first data row
        assert ws.cell(row=7, column=1).value == "S001"
        # Row 8 = second data row
        assert ws.cell(row=8, column=1).value == "S002"

    def test_q12a_format_uses_pipe_separator(self, mock_results, tmp_path):
        """Q12A should use pipe separator for multiple brands."""
        from process import build_client_format

        output_path = str(tmp_path / "client_output.xlsx")
        build_client_format(mock_results, output_path)

        wb = load_workbook(output_path)
        ws = wb.active

        q12a_val = ws.cell(row=7, column=2).value
        # Should contain both brands with pipe separator
        assert "MEVIUS" in q12a_val
        assert "ARA" in q12a_val
        assert "|" in q12a_val

    def test_q12b_format_uses_pipe_separator(self, mock_results, tmp_path):
        """Q12B SKUs should use pipe separator."""
        from process import build_client_format

        output_path = str(tmp_path / "client_output.xlsx")
        build_client_format(mock_results, output_path)

        wb = load_workbook(output_path)
        ws = wb.active

        q12b_val = ws.cell(row=7, column=3).value
        assert "MEVIUS ORIGINAL" in q12b_val
        assert "ARA RED" in q12b_val
        assert "|" in q12b_val


# ═══════════════════════════════════════════════════════════════════════════════
# 7. test_correction_loop_improves_prompt
# ═══════════════════════════════════════════════════════════════════════════════


class TestCorrectionLoopImprovesPrompt:
    def test_saved_correction_appears_in_prompt(self, tmp_path):
        """Save a correction, then verify format_corrections_for_prompt includes it."""
        from corrections import save_correction, load_corrections, format_corrections_for_prompt

        path = str(tmp_path / "corrections.json")

        correction = {
            "serial": "T001",
            "ai_result": {"brands": ["MEVIUS"], "skus": ["MEVIUS ORIGINAL"]},
            "corrected_result": {"brands": ["MEVIUS", "ESSE"], "skus": ["MEVIUS ORIGINAL", "ESSE CHANGE"]},
            "notes": "Missed ESSE pack behind glass glare",
        }
        save_correction(correction, path=path)

        loaded = load_corrections(path=path)
        text = format_corrections_for_prompt(loaded)

        assert "LEARNING FROM PAST CORRECTIONS" in text
        assert "MEVIUS" in text
        assert "ESSE" in text
        assert "ESSE CHANGE" in text
        assert "Missed ESSE pack behind glass glare" in text
        assert "brand_added" in text  # correction type

    def test_multiple_corrections_all_included(self, tmp_path):
        """Multiple corrections should all appear in the formatted prompt."""
        from corrections import save_correction, load_corrections, format_corrections_for_prompt

        path = str(tmp_path / "corrections.json")

        save_correction({
            "serial": "T001",
            "ai_result": {"brands": ["MEVIUS"], "skus": []},
            "corrected_result": {"brands": ["ARA"], "skus": []},
            "notes": "Wrong brand",
        }, path=path)

        save_correction({
            "serial": "T002",
            "ai_result": {"brands": ["555"], "skus": ["555 GOLD"]},
            "corrected_result": {"brands": ["555"], "skus": ["555 RED"]},
            "notes": "Wrong SKU color",
        }, path=path)

        loaded = load_corrections(path=path)
        text = format_corrections_for_prompt(loaded)

        assert "Correction 1" in text
        assert "Correction 2" in text
        assert "Wrong brand" in text
        assert "Wrong SKU color" in text


# ═══════════════════════════════════════════════════════════════════════════════
# 8. test_confidence_scoring_factors
# ═══════════════════════════════════════════════════════════════════════════════


class TestConfidenceScoringFactors:
    @patch("confidence.load_corrections", return_value=[])
    def test_perfect_conditions_score_high(self, mock_corr):
        """Perfect conditions should yield a high score."""
        from confidence import compute_confidence

        result = compute_confidence(
            ai_confidence="high",
            brands_found=["MEVIUS", "ARA"],
            skus_found=["MEVIUS ORIGINAL", "ARA RED"],
            unidentified_packs=0,
            num_images=1,
        )
        assert result["score"] >= 80
        assert result["level"] == "high"

    @patch("confidence.load_corrections", return_value=[])
    def test_low_ai_confidence_lowers_score(self, mock_corr):
        """Low AI confidence should significantly reduce the score."""
        from confidence import compute_confidence

        high = compute_confidence(
            ai_confidence="high",
            brands_found=["MEVIUS"],
            skus_found=["MEVIUS ORIGINAL"],
            unidentified_packs=0,
            num_images=1,
        )
        low = compute_confidence(
            ai_confidence="low",
            brands_found=["MEVIUS"],
            skus_found=["MEVIUS ORIGINAL"],
            unidentified_packs=0,
            num_images=1,
        )
        assert low["score"] < high["score"]
        # The ai_confidence factor should be documented
        assert low["factors"]["ai_confidence"]["value"] == "low"

    @patch("confidence.load_corrections", return_value=[])
    def test_missing_skus_lowers_score(self, mock_corr):
        """Having brands without SKUs should reduce confidence."""
        from confidence import compute_confidence

        with_skus = compute_confidence(
            ai_confidence="high",
            brands_found=["MEVIUS", "ARA"],
            skus_found=["MEVIUS ORIGINAL", "ARA RED"],
            unidentified_packs=0,
            num_images=1,
        )
        without_skus = compute_confidence(
            ai_confidence="high",
            brands_found=["MEVIUS", "ARA"],
            skus_found=[],
            unidentified_packs=0,
            num_images=1,
        )
        assert without_skus["score"] < with_skus["score"]

    @patch("confidence.load_corrections", return_value=[])
    def test_many_unidentified_packs_drops_score(self, mock_corr):
        """Many unidentified packs should significantly drop the score."""
        from confidence import compute_confidence

        result = compute_confidence(
            ai_confidence="high",
            brands_found=["MEVIUS"],
            skus_found=["MEVIUS ORIGINAL"],
            unidentified_packs=5,
            num_images=1,
        )
        assert result["score"] < 90
        assert result["factors"]["unidentified_packs"]["penalty"] > 0

    @patch("confidence.load_corrections", return_value=[])
    def test_inconsistent_multi_image_drops_score(self, mock_corr):
        """Images seeing completely different brands should lower confidence."""
        from confidence import compute_confidence

        result = compute_confidence(
            ai_confidence="high",
            brands_found=["MEVIUS", "ARA", "555"],
            skus_found=["MEVIUS ORIGINAL", "ARA RED", "555 GOLD"],
            unidentified_packs=0,
            num_images=3,
            brands_per_image=[["MEVIUS"], ["ARA"], ["555"]],
        )
        # No overlap between any images = 0% consistency
        assert result["factors"]["multi_image_consistency"]["score"] == 0
        assert result["score"] < 90

    @patch("confidence.load_corrections", return_value=[])
    def test_score_always_in_valid_range(self, mock_corr):
        """Score should always be between 0 and 100 regardless of inputs."""
        from confidence import compute_confidence

        # Worst case scenario
        result = compute_confidence(
            ai_confidence="low",
            brands_found=["MEVIUS"],
            skus_found=[],
            unidentified_packs=20,
            num_images=3,
            brands_per_image=[["MEVIUS"], ["ARA"], ["555"]],
        )
        assert 0 <= result["score"] <= 100


# ═══════════════════════════════════════════════════════════════════════════════
# 9. test_rate_limiter_with_multiple_threads
# ═══════════════════════════════════════════════════════════════════════════════


class TestRateLimiterMultiThread:
    def test_threads_are_properly_spaced(self):
        """Spawn 5 threads that call limiter.wait(), verify proper spacing."""
        from rate_limiter import RateLimiter

        # Use a high RPM so the test doesn't take too long
        # but low enough that spacing is measurable
        limiter = RateLimiter("unknown-test-model")
        # Override to a controlled interval
        limiter.rpm = 120  # 2 per second
        limiter.interval = 0.5  # 500ms between requests

        timestamps = []
        lock = threading.Lock()

        def worker():
            limiter.wait()
            with lock:
                timestamps.append(time.time())

        threads = [threading.Thread(target=worker) for _ in range(5)]
        start = time.time()
        for t in threads:
            t.start()
        for t in threads:
            t.join(timeout=10)

        # Should have exactly 5 timestamps
        assert len(timestamps) == 5

        # Sort timestamps and check spacing
        timestamps.sort()
        for i in range(1, len(timestamps)):
            gap = timestamps[i] - timestamps[i - 1]
            # Each gap should be at least ~0.5s (with some tolerance)
            assert gap >= 0.4, f"Gap between request {i-1} and {i} was only {gap:.3f}s"

        # Total time should be at least (5-1) * 0.5 = 2.0 seconds
        total = timestamps[-1] - timestamps[0]
        assert total >= 1.6  # with tolerance


# ═══════════════════════════════════════════════════════════════════════════════
# 10. test_cost_tracking_accumulates
# ═══════════════════════════════════════════════════════════════════════════════


class TestCostTrackingAccumulates:
    @pytest.fixture(autouse=True)
    def _use_temp_db(self, tmp_path, monkeypatch):
        """Point database module to a temporary SQLite DB for test isolation."""
        import database
        db_path = tmp_path / "test_chhat.db"
        monkeypatch.setattr(database, "DB_PATH", db_path)
        # Reset the cached connection so it reconnects to the temp DB
        database._connection = None

    def test_multiple_calls_accumulate(self):
        """Track multiple calls for a job, verify totals are correct."""
        from cost_tracker import track_call, get_job_cost

        job_id = "test-job-001"
        track_call(job_id, "gemini-2.5-pro", "analysis")
        track_call(job_id, "gemini-2.5-flash", "ocr")
        track_call(job_id, "gemini-2.5-pro", "sku_refinement")

        result = get_job_cost(job_id)
        assert result["total_calls"] == 3
        # Cost = 0.005 + 0.001 + 0.005 = 0.011
        assert abs(result["total_cost"] - 0.011) < 0.0001
        assert result["by_type"]["analysis"] == 1
        assert result["by_type"]["ocr"] == 1
        assert result["by_type"]["sku_refinement"] == 1

    def test_separate_jobs_tracked_independently(self):
        """Different job IDs should have independent cost tracking."""
        from cost_tracker import track_call, get_job_cost

        track_call("job-A", "gemini-2.5-pro", "analysis")
        track_call("job-A", "gemini-2.5-pro", "analysis")
        track_call("job-B", "claude-sonnet-4-6", "analysis")

        job_a = get_job_cost("job-A")
        job_b = get_job_cost("job-B")

        assert job_a["total_calls"] == 2
        assert job_b["total_calls"] == 1
        assert abs(job_a["total_cost"] - 0.010) < 0.0001
        assert abs(job_b["total_cost"] - 0.013) < 0.0001

    def test_total_cost_aggregates_all_jobs(self):
        """get_total_cost should aggregate across all jobs."""
        from cost_tracker import track_call, get_total_cost

        track_call("job-X", "gemini-2.5-pro", "analysis")
        track_call("job-Y", "gemini-2.5-pro", "analysis")

        total = get_total_cost()
        assert total["total_calls"] == 2
        assert abs(total["total_cost"] - 0.010) < 0.0001
        assert "gemini-2.5-pro" in total["by_model"]

    def test_nonexistent_job_returns_zeros(self):
        """Getting cost for a non-existent job should return zeros."""
        from cost_tracker import get_job_cost

        result = get_job_cost("nonexistent")
        assert result["total_calls"] == 0
        assert result["total_cost"] == 0.0
