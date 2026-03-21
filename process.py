#!/usr/bin/env python3
"""
Process the client's CHHAT Excel file:
1. Read Raw Data sheet (serial + image links)
2. Fetch and analyze images in PARALLEL using thread pool
3. Output Excel with: serial | embedded images | Q12A brands | Q12B SKUs | brand count
"""

import argparse
import io
import os
import sys
import time
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed

from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

from image_analyzer import fetch_image, analyze_image, _resize_image, get_available_models
from brands import format_q12a, BRANDS_AND_SKUS
from corrections import find_relevant_corrections, format_corrections_for_prompt, get_correction_stats
from confidence import compute_confidence
from rate_limiter import RateLimiter


def parse_args():
    parser = argparse.ArgumentParser(description="Process CHHAT cigarette survey images.")
    parser.add_argument("input_file", help="Path to the CHHAT Excel file")
    parser.add_argument("--output", "-o", help="Output file path (default: <input>_results.xlsx)")
    parser.add_argument("--model", "-m", default="gemini-2.5-pro",
                        help=f"AI model. Options: {', '.join(get_available_models())}")
    parser.add_argument("--delay", type=float, default=0.5, help="Delay between API calls (seconds)")
    parser.add_argument("--start-row", type=int, default=3, help="First data row in Raw Data sheet (default: 3)")
    parser.add_argument("--limit", type=int, help="Only process first N rows (for testing)")
    parser.add_argument("--dry-run", action="store_true", help="Show what would be processed")
    parser.add_argument("--photo-cols", default="B,C,D", help="Columns with Q32 photo links (default: B,C,D)")
    parser.add_argument("--workers", "-w", type=int, default=5, help="Number of parallel workers (default: 5)")
    return parser.parse_args()


def read_raw_data(file_path, start_row=3, photo_cols_str="B,C,D"):
    """Read the Raw Data sheet and extract serial numbers + image URLs."""
    from openpyxl import load_workbook
    from openpyxl.utils import column_index_from_string

    wb = load_workbook(file_path)
    ws = wb["Raw data "]

    photo_cols = [column_index_from_string(c.strip()) for c in photo_cols_str.split(",")]

    rows = []
    for row_num in range(start_row, ws.max_row + 1):
        serial = ws.cell(row=row_num, column=1).value
        if not serial:
            continue

        urls = []
        for col in photo_cols:
            val = ws.cell(row=row_num, column=col).value
            if val and str(val).startswith("http"):
                urls.append(str(val).strip())

        rows.append({"row_number": row_num, "serial": serial, "urls": urls})

    return rows


def create_thumbnail(image_data: bytes, max_size=(200, 200)) -> bytes:
    """Create a thumbnail for embedding in Excel."""
    img = PILImage.open(io.BytesIO(image_data))
    if img.mode in ("RGBA", "P", "LA"):
        img = img.convert("RGB")
    img.thumbnail(max_size, PILImage.Resampling.LANCZOS)
    buf = io.BytesIO()
    img.save(buf, format="JPEG", quality=75)
    return buf.getvalue()


def _process_single_image(url, model, api_keys, correction_context):
    """Process a single image — used by both sequential and parallel modes."""
    image_data, media_type = fetch_image(url)
    thumbnail = create_thumbnail(image_data)
    analysis = analyze_image(
        image_data, media_type,
        model=model, api_keys=api_keys,
        correction_context=correction_context,
    )
    return {
        "url": url,
        "thumbnail": thumbnail,
        "analysis": analysis,
    }


def _process_outlet(row, model, api_keys, correction_context, limiter=None):
    """Process all images for a single outlet — runs in thread pool."""
    serial = row["serial"]
    urls = row["urls"]

    all_brands = set()
    all_skus = set()
    thumbnails = []
    total_unidentified = 0
    brands_per_image = []
    ai_confidences = []
    error = None

    for url in urls:
        try:
            if limiter:
                limiter.wait()
            result = _process_single_image(url, model, api_keys, correction_context)
            thumbnails.append(result["thumbnail"])
            analysis = result["analysis"]

            if "error" not in analysis:
                img_brands = []
                for brand_entry in analysis.get("brands_found", []):
                    brand_name = brand_entry.get("brand", "")
                    if brand_name in BRANDS_AND_SKUS:
                        all_brands.add(brand_name)
                        img_brands.append(brand_name)
                    for sku in brand_entry.get("skus", []):
                        if sku:
                            all_skus.add(sku)
                brands_per_image.append(img_brands)
                total_unidentified += analysis.get("unidentified_packs", 0)
                ai_confidences.append(analysis.get("confidence", "medium"))
            else:
                error = analysis["error"]

        except Exception as e:
            error = str(e)

    brands_sorted = sorted(all_brands)
    skus_sorted = sorted(all_skus)

    # Compute confidence score
    worst_ai_conf = "high"
    conf_rank = {"low": 0, "medium": 1, "high": 2}
    for c in ai_confidences:
        if conf_rank.get(c, 1) < conf_rank.get(worst_ai_conf, 2):
            worst_ai_conf = c

    confidence_result = compute_confidence(
        ai_confidence=worst_ai_conf,
        brands_found=brands_sorted,
        skus_found=skus_sorted,
        unidentified_packs=total_unidentified,
        num_images=len(urls),
        brands_per_image=brands_per_image,
    )

    return {
        "serial": serial,
        "brands": brands_sorted,
        "skus": skus_sorted,
        "thumbnails": thumbnails,
        "unidentified_packs": total_unidentified,
        "confidence": confidence_result["level"],
        "confidence_score": confidence_result["score"],
        "confidence_factors": confidence_result["factors"],
        "error": error if not brands_sorted and error else None,
    }


def build_output(results: list[dict], output_path: str):
    """Build the output Excel with embedded images and brand analysis."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    # Headers
    headers = [
        ("A", "Serial Number", 16),
        ("B", "Q32 Image 1", 30),
        ("C", "Q32 Image 2", 30),
        ("D", "Q32 Image 3", 30),
        ("E", "Q12A - Brands", 50),
        ("F", "Q12B - SKUs", 60),
        ("G", "Brand Count", 14),
        ("H", "Unidentified Packs", 18),
        ("I", "Confidence", 18),
        ("J", "Status", 12),
    ]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")

    for col_letter, title, width in headers:
        cell = ws[f"{col_letter}1"]
        cell.value = title
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[col_letter].width = width

    # Data rows
    for i, result in enumerate(results):
        row = i + 2
        ws.row_dimensions[row].height = 120

        ws.cell(row=row, column=1, value=result["serial"])
        ws.cell(row=row, column=1).alignment = Alignment(vertical="center")

        # Embed images
        thumbnails = result.get("thumbnails", [])
        for img_idx, thumb_data in enumerate(thumbnails[:3]):
            col = img_idx + 2
            try:
                img_stream = io.BytesIO(thumb_data)
                img = XlImage(img_stream)
                img.width = 180
                img.height = 110
                cell_ref = f"{get_column_letter(col)}{row}"
                ws.add_image(img, cell_ref)
            except Exception:
                ws.cell(row=row, column=col, value="[image error]")

        # Q12A - Brands
        brands = result.get("brands", [])
        ws.cell(row=row, column=5, value=format_q12a(brands))
        ws.cell(row=row, column=5).alignment = Alignment(wrap_text=True, vertical="center")

        # Q12B - SKUs
        skus = result.get("skus", [])
        ws.cell(row=row, column=6, value=" | ".join(skus) if skus else "")
        ws.cell(row=row, column=6).alignment = Alignment(wrap_text=True, vertical="center")

        # Brand count
        ws.cell(row=row, column=7, value=len(brands))
        ws.cell(row=row, column=7).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=7).font = Font(bold=True, size=14)

        # Unidentified packs
        unidentified = result.get("unidentified_packs", 0)
        cell_unid = ws.cell(row=row, column=8, value=unidentified)
        cell_unid.alignment = Alignment(horizontal="center", vertical="center")
        if unidentified and unidentified > 0:
            cell_unid.font = Font(color="FF8C00", bold=True)

        # Confidence (show score + level)
        confidence = result.get("confidence", "")
        conf_score = result.get("confidence_score", "")
        conf_display = f"{confidence} ({conf_score}%)" if conf_score else confidence
        cell_conf = ws.cell(row=row, column=9, value=conf_display)
        cell_conf.alignment = Alignment(horizontal="center", vertical="center")
        if confidence == "high":
            cell_conf.font = Font(color="00B050", bold=True)
        elif confidence == "medium":
            cell_conf.font = Font(color="FF8C00", bold=True)
        elif confidence == "low":
            cell_conf.font = Font(color="FF0000", bold=True)

        # Status
        if result.get("error"):
            status_cell = ws.cell(row=row, column=10, value="ERROR")
            status_cell.font = Font(color="FF0000", bold=True)
        else:
            status_cell = ws.cell(row=row, column=10, value="OK")
            status_cell.font = Font(color="00B050", bold=True)
        status_cell.alignment = Alignment(horizontal="center", vertical="center")

    wb.save(output_path)
    return output_path


def build_client_format(results: list[dict], output_path: str):
    """
    Build output in the client's Result Format (Sheet 3 format):
    Row 5: Respondent.Serial | Q12A | Q12B
    Row 6: Serial number | PLEASE SELECT THE BRAND... | PLEASE SELECT THE SKU...
    Row 7+: data
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Result format "

    header_font = Font(bold=True, size=11)

    # Row 5: field codes
    ws.cell(row=5, column=1, value="Respondent.Serial").font = header_font
    ws.cell(row=5, column=2, value="Q12A").font = header_font
    ws.cell(row=5, column=3, value="Q12B").font = header_font

    # Row 6: descriptions
    ws.cell(row=6, column=1, value="Serial number").font = header_font
    ws.cell(row=6, column=2, value="PLEASE SELECT THE BRAND OF TOBACCO SELLING IN THE OUTLET").font = header_font
    ws.cell(row=6, column=3, value="PLEASE SELECT THE SKU AVAILABLE IN THE OUTLET").font = header_font

    # Column widths
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 80
    ws.column_dimensions["C"].width = 80

    # Data rows
    for i, result in enumerate(results):
        row = i + 7
        ws.cell(row=row, column=1, value=result["serial"])

        brands = result.get("brands", [])
        skus = result.get("skus", [])

        # Q12A format: BRAND_Khmer | BRAND_Khmer
        q12a = format_q12a(brands)
        ws.cell(row=row, column=2, value=q12a)
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True)

        # Q12B format: SKU1 | SKU2
        q12b = " | ".join(skus) if skus else ""
        ws.cell(row=row, column=3, value=q12b)
        ws.cell(row=row, column=3).alignment = Alignment(wrap_text=True)

    wb.save(output_path)
    return output_path


def main():
    load_dotenv()
    args = parse_args()

    input_path = Path(args.input_file)
    if not input_path.exists():
        print(f"Error: File not found: {input_path}")
        sys.exit(1)

    output_path = args.output or str(input_path.with_stem(input_path.stem + "_results"))

    api_keys = {
        "claude": os.getenv("ANTHROPIC_API_KEY", ""),
        "gemini": os.getenv("GEMINI_API_KEY", ""),
        "fireworks": os.getenv("FIREWORKS_API_KEY", ""),
    }

    rows = read_raw_data(str(input_path), start_row=args.start_row, photo_cols_str=args.photo_cols)

    if args.limit:
        rows = rows[:args.limit]

    print(f"\n{'='*60}")
    print(f"  CHHAT Cigarette Brand Analyzer")
    print(f"{'='*60}")
    print(f"  Input:   {input_path}")
    print(f"  Output:  {output_path}")
    print(f"  Rows:    {len(rows)}")
    print(f"  Model:   {args.model}")
    print(f"  Workers: {args.workers} (parallel)")

    if args.dry_run:
        print(f"\n--- DRY RUN ---")
        for row in rows:
            print(f"  Serial {row['serial']}: {len(row['urls'])} image(s)")
            for u in row["urls"]:
                print(f"    {u[:80]}...")
        sys.exit(0)

    # Load past corrections
    stats = get_correction_stats()
    if stats["total"] > 0:
        print(f"  Corrections: {stats['total']} past corrections loaded")
        recent_corrections = find_relevant_corrections(limit=5)
        correction_context = format_corrections_for_prompt(recent_corrections)
    else:
        print(f"  Corrections: none yet")
        correction_context = ""

    total_images = sum(len(r["urls"]) for r in rows)
    print(f"  Total images: {total_images}")

    print(f"\n{'─'*60}")
    print(f"  Processing {len(rows)} outlets ({args.workers} parallel workers)...\n")

    start_time = time.time()

    # ── Parallel processing ──────────────────────────────────────────
    results = [None] * len(rows)
    completed = 0

    limiter = RateLimiter(args.model)
    safe_workers = min(args.workers, limiter.get_safe_workers())
    if safe_workers < args.workers:
        print(f"  Rate limit: {limiter.rpm} RPM for {args.model}, capping workers at {safe_workers}")

    with ThreadPoolExecutor(max_workers=safe_workers) as executor:
        future_to_idx = {}
        for idx, row in enumerate(rows):
            future = executor.submit(
                _process_outlet, row, args.model, api_keys, correction_context, limiter
            )
            future_to_idx[future] = idx

        for future in as_completed(future_to_idx):
            idx = future_to_idx[future]
            completed += 1
            try:
                result = future.result()
                results[idx] = result
                serial = result["serial"]
                brands = result["brands"]
                conf = result.get("confidence", "?")
                conf_score = result.get("confidence_score", "?")
                unid = result.get("unidentified_packs", 0)
                unid_str = f" | Unidentified: {unid}" if unid else ""
                print(
                    f"  [{completed}/{len(rows)}] Serial {serial}: "
                    f"{', '.join(brands) if brands else 'None'} ({len(brands)}) "
                    f"| Confidence: {conf} ({conf_score}%){unid_str}"
                )
            except Exception as e:
                row = rows[idx]
                results[idx] = {
                    "serial": row["serial"],
                    "brands": [],
                    "skus": [],
                    "thumbnails": [],
                    "unidentified_packs": 0,
                    "confidence": "low",
                    "confidence_score": 0,
                    "error": str(e),
                }
                print(f"  [{completed}/{len(rows)}] Serial {row['serial']}: ERROR — {e}")

    elapsed = time.time() - start_time

    # Build output
    print(f"\n{'─'*60}")
    print(f"  Writing results to: {output_path}")
    build_output(results, output_path)

    # Summary
    total_brands = set()
    for r in results:
        if r:
            total_brands.update(r.get("brands", []))

    avg_conf = sum(r.get("confidence_score", 0) for r in results if r) / len(results) if results else 0

    print(f"\n{'='*60}")
    print(f"  DONE in {elapsed:.1f}s ({elapsed/len(rows):.1f}s per outlet)")
    print(f"  Outlets processed: {len(results)}")
    print(f"  Unique brands: {len(total_brands)}")
    print(f"  Avg confidence: {avg_conf:.0f}%")
    if total_brands:
        print(f"  Brands: {', '.join(sorted(total_brands))}")
    print(f"  Output: {output_path}")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
