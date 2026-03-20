#!/usr/bin/env python3
"""Generate a demo Excel file with real cigarette image URLs."""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def create_demo():
    wb = Workbook()
    ws = wb.active
    ws.title = "Inspection Data"

    # Headers
    headers = ["Inspection ID", "Store Location", "Image Link", "Inspector", "Date", "Notes"]
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Real image URLs from Wikimedia Commons
    rows = [
        [
            "INS-001",
            "Downtown Convenience Store",
            "https://upload.wikimedia.org/wikipedia/commons/8/8c/Marlboro_cigarettes.jpg",
            "Sarah Chen",
            "2026-03-10",
            "Front counter display",
        ],
        [
            "INS-002",
            "Airport Duty Free Shop",
            "https://upload.wikimedia.org/wikipedia/commons/e/e9/Chesterfield_Blue_German.jpg",
            "James Rivera",
            "2026-03-11",
            "Terminal B kiosk",
        ],
        [
            "INS-003",
            "Gas Station - Highway 101",
            "https://upload.wikimedia.org/wikipedia/commons/2/27/Captain_Black_Blue_Cigarette_Pack_03.jpg",
            "Sarah Chen",
            "2026-03-12",
            "Behind register area",
        ],
        [
            "INS-004",
            "Corner Store - 5th Ave",
            "https://upload.wikimedia.org/wikipedia/commons/1/15/Cigarette_LD_%283%29.jpg",
            "Michael Park",
            "2026-03-13",
            "Wall display unit",
        ],
        [
            "INS-005",
            "Supermarket - Oak Mall",
            "https://upload.wikimedia.org/wikipedia/commons/a/ab/Peace_cigarettes_pack%2C_front.JPG",
            "James Rivera",
            "2026-03-14",
            "Customer service counter",
        ],
        [
            "INS-006",
            "Tobacco Shop - Main St",
            "https://upload.wikimedia.org/wikipedia/commons/4/43/Time_cigarette_pack.jpg",
            "Michael Park",
            "2026-03-15",
            "Main display case",
        ],
        [
            "INS-007",
            "Liquor Store - Pine Rd",
            "https://upload.wikimedia.org/wikipedia/commons/a/ac/Sultan_cigarettes_pack%2C.JPG",
            "Sarah Chen",
            "2026-03-16",
            "Shelf behind counter",
        ],
        [
            "INS-008",
            "Hotel Gift Shop",
            "https://upload.wikimedia.org/wikipedia/commons/9/95/Condal_Cigarette_Pack%2C_1941.jpg",
            "James Rivera",
            "2026-03-17",
            "Lobby shop vitrine",
        ],
    ]

    for row_data in rows:
        ws.append(row_data)

    # Style data rows
    for row in range(2, len(rows) + 2):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    # Column widths
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 55
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 28

    output = "demo_input.xlsx"
    wb.save(output)
    print(f"Created {output} with {len(rows)} inspection records.")
    print("Each row has a real Wikimedia Commons image URL of cigarette packs.")


if __name__ == "__main__":
    create_demo()
