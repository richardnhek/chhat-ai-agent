#!/usr/bin/env python3
"""Generate a sample Excel file for testing the analyzer."""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill


def create_sample():
    wb = Workbook()
    ws = wb.active
    ws.title = "Cigarette Inventory"

    # Headers
    headers = ["ID", "Location", "Image Link", "Inspector", "Date"]
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")

    for col, title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Sample data — replace these URLs with real image URLs for testing
    sample_rows = [
        [1, "Store A - Front Counter", "https://example.com/image1.jpg", "John", "2026-03-15"],
        [2, "Store A - Back Shelf", "https://example.com/image2.jpg", "John", "2026-03-15"],
        [3, "Store B - Display Case", "https://example.com/image3.png", "Sarah", "2026-03-16"],
        [4, "Store B - Register Area", "https://example.com/image4.jpg", "Sarah", "2026-03-16"],
        [5, "Store C - Main Floor", "https://example.com/image5.jpg", "Mike", "2026-03-17"],
    ]

    for row_data in sample_rows:
        ws.append(row_data)

    # Adjust column widths
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15

    output = "sample_input.xlsx"
    wb.save(output)
    print(f"Created {output} with {len(sample_rows)} sample rows.")
    print("Replace the example.com URLs with real image URLs before running the analyzer.")


if __name__ == "__main__":
    create_sample()
