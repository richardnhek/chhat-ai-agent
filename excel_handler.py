"""
Excel handler — reads links from an Excel sheet and writes analysis results back.
"""

import re
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill


URL_PATTERN = re.compile(
    r"https?://[^\s<>\"']+\.(?:jpg|jpeg|png|gif|bmp|webp|tiff|svg)",
    re.IGNORECASE,
)

# Broader pattern for any URL (fallback if no image-specific URL found)
GENERAL_URL_PATTERN = re.compile(r"https?://[^\s<>\"']+", re.IGNORECASE)


def _looks_like_url(value: str) -> bool:
    """Check if a cell value looks like a URL."""
    if not isinstance(value, str):
        return False
    return bool(GENERAL_URL_PATTERN.match(value.strip()))


def _detect_link_column(ws) -> int | None:
    """
    Auto-detect which column contains image links by scanning the first
    few data rows and checking for URL patterns or hyperlinks.
    Returns the 1-based column index, or None if not found.
    """
    max_scan_rows = min(20, ws.max_row)
    col_url_counts: dict[int, int] = {}

    for row in range(2, max_scan_rows + 1):  # skip header
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            # Check for hyperlinks
            if cell.hyperlink and cell.hyperlink.target:
                col_url_counts[col] = col_url_counts.get(col, 0) + 1
                continue
            # Check cell value
            val = cell.value
            if val and _looks_like_url(str(val)):
                col_url_counts[col] = col_url_counts.get(col, 0) + 1

    if not col_url_counts:
        return None
    return max(col_url_counts, key=col_url_counts.get)


def _extract_url(cell) -> str | None:
    """Extract a URL from a cell (hyperlink or cell value)."""
    if cell.hyperlink and cell.hyperlink.target:
        return cell.hyperlink.target.strip()
    val = cell.value
    if val and _looks_like_url(str(val)):
        return str(val).strip()
    return None


def load_excel(
    file_path: str,
    sheet_name: str | None = None,
    link_column: int | str | None = None,
):
    """
    Load an Excel file and return structured row data with URLs.

    Args:
        file_path: Path to the .xlsx file.
        sheet_name: Name of the sheet to read (default: active sheet).
        link_column: Column containing links. Can be:
            - An int (1-based column index)
            - A str (column letter like "B" or column header name)
            - None (auto-detect)

    Returns:
        dict with keys:
            - workbook: the openpyxl Workbook object
            - worksheet: the active Worksheet
            - headers: list of header values
            - rows: list of dicts with {row_number, url, header_values}
            - link_col_index: 1-based index of the link column
    """
    wb = load_workbook(file_path)
    ws = wb[sheet_name] if sheet_name else wb.active

    # Read headers from row 1
    headers = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        headers.append(str(val) if val else f"Column_{col}")

    # Resolve link column
    link_col_idx = None
    if link_column is None:
        link_col_idx = _detect_link_column(ws)
    elif isinstance(link_column, int):
        link_col_idx = link_column
    elif isinstance(link_column, str):
        # Try as column letter first (A, B, C, ...)
        if len(link_column) <= 3 and link_column.isalpha():
            from openpyxl.utils import column_index_from_string
            link_col_idx = column_index_from_string(link_column)
        else:
            # Try as header name
            for i, h in enumerate(headers):
                if h.lower().strip() == link_column.lower().strip():
                    link_col_idx = i + 1
                    break

    if link_col_idx is None:
        raise ValueError(
            "Could not detect the link column. "
            "Please specify it with --link-column (column letter or header name)."
        )

    # Extract rows
    rows = []
    for row_num in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_num, column=link_col_idx)
        url = _extract_url(cell)

        # Skip completely empty rows
        row_values = {}
        has_data = False
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=row_num, column=col).value
            row_values[headers[col - 1]] = val
            if val is not None:
                has_data = True

        if not has_data:
            continue

        rows.append({
            "row_number": row_num,
            "url": url,
            "values": row_values,
        })

    return {
        "workbook": wb,
        "worksheet": ws,
        "headers": headers,
        "rows": rows,
        "link_col_index": link_col_idx,
    }


def write_results(
    workbook,
    worksheet,
    results: list[dict],
    output_path: str,
):
    """
    Write analysis results back to the Excel file in new columns.

    Args:
        workbook: openpyxl Workbook
        worksheet: openpyxl Worksheet
        results: list of dicts with {row_number, cigarette_count, brands, raw_analysis, error}
        output_path: where to save the updated file
    """
    # Find the next available columns for output
    next_col = worksheet.max_column + 1
    count_col = next_col
    brands_col = next_col + 1
    details_col = next_col + 2
    status_col = next_col + 3

    # Style for headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", wrap_text=True)

    breakdown_col = next_col + 4
    est_cigs_col = next_col + 5

    # Write headers
    for col, title in [
        (count_col, "Total Packs"),
        (brands_col, "Brands Identified"),
        (details_col, "Brand Breakdown"),
        (status_col, "Status"),
        (breakdown_col, "Visibility Notes"),
        (est_cigs_col, "Est. Cigarettes"),
    ]:
        cell = worksheet.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Write results
    for result in results:
        row = result["row_number"]

        if result.get("error"):
            worksheet.cell(row=row, column=count_col, value="N/A")
            worksheet.cell(row=row, column=brands_col, value="N/A")
            worksheet.cell(row=row, column=details_col, value=result["error"])
            status_cell = worksheet.cell(row=row, column=status_col, value="ERROR")
            status_cell.font = Font(color="FF0000", bold=True)
        else:
            worksheet.cell(row=row, column=count_col, value=result.get("total_packs", "N/A"))
            worksheet.cell(row=row, column=brands_col, value=result.get("brands", "N/A"))
            worksheet.cell(row=row, column=details_col, value=result.get("brand_breakdown", ""))
            status_cell = worksheet.cell(row=row, column=status_col, value="OK")
            status_cell.font = Font(color="00B050", bold=True)
            worksheet.cell(row=row, column=breakdown_col, value=result.get("visibility_notes", ""))
            worksheet.cell(row=row, column=est_cigs_col, value=result.get("estimated_cigarettes", "N/A"))

    # Auto-adjust column widths
    for col in [count_col, brands_col, details_col, status_col]:
        max_len = 0
        for row in range(1, worksheet.max_row + 1):
            val = worksheet.cell(row=row, column=col).value
            if val:
                max_len = max(max_len, len(str(val)))
        worksheet.column_dimensions[
            worksheet.cell(row=1, column=col).column_letter
        ].width = min(max_len + 4, 60)

    workbook.save(output_path)
    return output_path
